from openpyxl import Workbook
from openpyxl.styles import Font
import datetime
import sys


class GenerateExcel():
    """Generates an Excel file"""
    def __init__(self, filename):
        self.filename = filename
        """Inits an Excle workbook
            ARG: 
                filname (str): filename (w/o ending)
        """
        self.wb = Workbook()
        # grab the active worksheet
        self.ws = self.wb.active
        self.store()

    def add_header(self):
        """Adds some header information to cells"""
        font = Font(size=16)
        self.ws.column_dimensions['A'].width = 30
        self.ws['A1'].font = font
        self.ws['A1'] = "Thermocycler Prüfbericht"
        self.ws['A2'] = "erzeugt von View.py (Thermocycler)"
        self.ws['A4'] = "Zeitstempel:"
        self.ws['B4'] = "Temp-1"
        self.ws['C4'] = "Temp-2"
        self.ws['D4'] = "Zyklen"


    def add_zell(self, position, value):
        """Adds values to cells
            ARG: 
                position (Str)        : cell position "A1"
                values   (Str..Int..) : any value 
            RETURN:
                None"""
        self.ws[position] =  value


    def store(self):
        """Save the file, handles PermissionErrors"""
        try:
            self.wb.save(self.filename + ".xlsx")
            print("File saved: " + str(self.filename + ".xlsx"))
        except PermissionError:
            print(sys.exc_info() [0])
            self.filename = input("File allread open? Enter a new filename:")
            self.store()
        except:
            print(sys.exc_info() [0])
            print("Unhandled Error file was not saved - sorry")


if __name__ == "__main__":

    generateExcel = GenerateExcel("sample")
    generateExcel.add_header()

    print("adding cells...")
    for i in range(6, 30):
        generateExcel.add_zell("A" + str(i), datetime.datetime.now())
        generateExcel.add_zell("B" + str(i), i + 2)

    generateExcel.store()






